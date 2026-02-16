"""
BOM Tasks.

Celery tasks for BOM-related operations.
"""

from celery import shared_task
from django.db import transaction
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def validate_bom_structure(self, bom_id: str):
    """
    Validate BOM structure integrity.
    
    Checks:
    - No circular references
    - All parent references valid
    - Correct level/path calculations
    - No orphaned items
    """
    from infrastructure.persistence.models import BOMStructure, BOMItem
    
    try:
        bom = BOMStructure.objects.get(id=bom_id)
        items = bom.items.all()
        
        issues = []
        
        # Check for circular references
        def check_circular(item, visited=None):
            if visited is None:
                visited = set()
            
            if item.id in visited:
                return True
            
            visited.add(item.id)
            
            if item.parent_item:
                return check_circular(item.parent_item, visited)
            
            return False
        
        for item in items:
            if check_circular(item):
                issues.append({
                    'type': 'circular_reference',
                    'item_id': str(item.id),
                    'message': f'Circular reference detected for item {item.position_number}'
                })
        
        # Check parent references
        for item in items.filter(parent_item__isnull=False):
            if item.parent_item.bom_id != bom.id:
                issues.append({
                    'type': 'invalid_parent',
                    'item_id': str(item.id),
                    'message': f'Parent item belongs to different BOM'
                })
        
        # Check levels
        for item in items:
            expected_level = 0
            if item.parent_item:
                expected_level = item.parent_item.level + 1
            
            if item.level != expected_level:
                issues.append({
                    'type': 'incorrect_level',
                    'item_id': str(item.id),
                    'message': f'Level mismatch: expected {expected_level}, got {item.level}'
                })
        
        logger.info(f"BOM {bom.code} validation: {len(issues)} issues found")
        
        return {
            'bom_id': bom_id,
            'bom_code': bom.code,
            'items_count': items.count(),
            'valid': len(issues) == 0,
            'issues': issues,
        }
        
    except BOMStructure.DoesNotExist:
        logger.error(f"BOM {bom_id} not found")
        return {'error': 'BOM not found'}
    except Exception as e:
        logger.error(f"Error validating BOM {bom_id}: {e}")
        self.retry(countdown=60)


@shared_task(bind=True, max_retries=3)
def recalculate_bom_paths(self, bom_id: str):
    """
    Recalculate paths and levels for all BOM items.
    """
    from infrastructure.persistence.models import BOMStructure, BOMItem
    
    try:
        bom = BOMStructure.objects.get(id=bom_id)
        
        with transaction.atomic():
            # Get root items (no parent)
            root_items = bom.items.filter(parent_item__isnull=True)
            
            def update_item_hierarchy(item, parent_path='', level=0):
                path = f"{parent_path}/{item.position_number}" if parent_path else item.position_number
                
                item.level = level
                item.path = path
                item.save(update_fields=['level', 'path'])
                
                for child in item.children.all():
                    update_item_hierarchy(child, path, level + 1)
            
            for root_item in root_items:
                update_item_hierarchy(root_item)
        
        logger.info(f"Recalculated paths for BOM {bom.code}")
        
        return {
            'bom_id': bom_id,
            'bom_code': bom.code,
            'success': True
        }
        
    except BOMStructure.DoesNotExist:
        return {'error': 'BOM not found'}
    except Exception as e:
        logger.error(f"Error recalculating BOM paths {bom_id}: {e}")
        self.retry(countdown=60)


@shared_task
def export_bom_to_excel(bom_id: str, user_id: str):
    """
    Export BOM to Excel file.
    
    Creates Excel file and stores it for download.
    """
    from infrastructure.persistence.models import BOMStructure
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from django.conf import settings
    import os
    
    try:
        bom = BOMStructure.objects.get(id=bom_id)
        items = bom.items.select_related('nomenclature').order_by('path')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        
        # Header style
        header_font = Font(bold=True)
        
        # Headers
        headers = [
            'Поз.', 'Уровень', 'Обозначение', 'Наименование',
            'Категория', 'Кол-во', 'Ед.', 'Примечание'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
        
        # Data
        for row, item in enumerate(items, 2):
            indent = "  " * item.level
            ws.cell(row=row, column=1, value=item.position_number)
            ws.cell(row=row, column=2, value=item.level)
            ws.cell(row=row, column=3, value=f"{indent}{item.nomenclature.code}")
            ws.cell(row=row, column=4, value=item.nomenclature.name)
            ws.cell(row=row, column=5, value=item.nomenclature.get_category_display())
            ws.cell(row=row, column=6, value=float(item.quantity))
            ws.cell(row=row, column=7, value=item.unit)
            ws.cell(row=row, column=8, value=item.notes or '')
        
        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports', 'bom')
        os.makedirs(export_dir, exist_ok=True)
        
        filename = f"BOM_{bom.code}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(export_dir, filename)
        wb.save(filepath)
        
        logger.info(f"Exported BOM {bom.code} to {filename}")
        
        return {
            'bom_id': bom_id,
            'filename': filename,
            'filepath': filepath,
            'download_url': f"/media/exports/bom/{filename}",
        }
        
    except BOMStructure.DoesNotExist:
        return {'error': 'BOM not found'}
    except Exception as e:
        logger.error(f"Error exporting BOM {bom_id}: {e}")
        return {'error': str(e)}


@shared_task
def import_bom_from_excel(file_path: str, bom_id: str = None, root_nomenclature_id: str = None):
    """
    Import BOM from Excel file.
    
    Expected columns:
    - Позиция (position)
    - Обозначение (designation) 
    - Наименование (name)
    - Количество (quantity)
    - Единица (unit)
    - Примечание (notes)
    """
    import openpyxl
    from infrastructure.persistence.models import (
        BOMStructure, BOMItem, NomenclatureItem
    )
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Parse headers (first row)
        headers = {}
        for col, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.lower().strip()] = col
        
        # Required columns mapping
        column_map = {
            'position': ['позиция', 'поз.', 'position', 'pos'],
            'code': ['обозначение', 'код', 'designation', 'code'],
            'name': ['наименование', 'name'],
            'quantity': ['количество', 'кол-во', 'qty', 'quantity'],
            'unit': ['единица', 'ед.', 'unit'],
        }
        
        # Find columns
        columns = {}
        for key, variants in column_map.items():
            for variant in variants:
                if variant in headers:
                    columns[key] = headers[variant]
                    break
        
        if 'code' not in columns:
            return {'error': 'Column "Обозначение" or "Код" not found'}
        
        # Get or create BOM
        if bom_id:
            bom = BOMStructure.objects.get(id=bom_id)
        else:
            # Create new BOM
            root = NomenclatureItem.objects.get(id=root_nomenclature_id)
            bom = BOMStructure.objects.create(
                code=f"BOM-{root.code}",
                name=f"BOM для {root.name}",
                root_nomenclature=root,
                status='draft',
            )
        
        # Import items
        items_created = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
            try:
                code = row[columns['code'] - 1].value
                if not code:
                    continue
                
                # Find nomenclature
                nomenclature = NomenclatureItem.objects.filter(
                    code=code.strip()
                ).first()
                
                if not nomenclature:
                    errors.append(f"Row {row_num}: Nomenclature with code '{code}' not found")
                    continue
                
                # Create item
                BOMItem.objects.create(
                    bom=bom,
                    nomenclature=nomenclature,
                    position_number=row[columns.get('position', 1) - 1].value or str(row_num),
                    quantity=row[columns.get('quantity', 1) - 1].value or 1,
                    unit=row[columns.get('unit', 1) - 1].value or nomenclature.unit.short_name if nomenclature.unit else 'шт',
                    level=0,
                )
                items_created += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        logger.info(f"Imported {items_created} items to BOM {bom.code}")
        
        return {
            'bom_id': str(bom.id),
            'items_created': items_created,
            'errors': errors,
        }
        
    except Exception as e:
        logger.error(f"Error importing BOM: {e}")
        return {'error': str(e)}

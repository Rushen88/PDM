"""
Catalog Views.

API views for nomenclature, suppliers, contractors, and catalog categories.
"""

from dataclasses import dataclass
from typing import Any
from uuid import uuid4

from rest_framework import status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import rest_framework as django_filters
from django.db import models
from django.db import transaction
from django.db.models import Count, Q, Exists, OuterRef

from openpyxl import load_workbook

from infrastructure.persistence.models import (
    CatalogCategory,
    NomenclatureItem,
    NomenclatureType,
    NomenclatureSupplier,
    NomenclatureCategoryChoices,
    Supplier,
    Contractor,
    ContactPerson,
    DelayReason,
    BOMStructure,
    BankDetails,
)
from ..serializers.catalog import (
    CatalogCategoryListSerializer,
    CatalogCategoryDetailSerializer,
    NomenclatureTypeSerializer,
    NomenclatureListSerializer,
    NomenclatureDetailSerializer,
    NomenclatureTreeSerializer,
    NomenclatureCategorySerializer,
    NomenclatureSupplierSerializer,
    SupplierListSerializer,
    SupplierDetailSerializer,
    ContractorListSerializer,
    ContractorDetailSerializer,
    ContactPersonSerializer,
    ContactPersonCreateSerializer,
    DelayReasonSerializer,
    BankDetailsSerializer,
    BankDetailsInlineSerializer,
)
from .base import BaseModelViewSet, BulkActionMixin


EXCEL_COLS = {
    'A': 1,  # catalog category name
    'B': 2,  # item name
    'C': 3,  # drawing number (manufactured only)
    'D': 4,  # unit
    'E': 5,  # description
    'F': 6,  # specifications
    'G': 7,  # nomenclature type name (purchased only)
}


def _to_str(v: Any) -> str:
    if v is None:
        return ''
    s = str(v).strip()
    return s


@dataclass(frozen=True)
class ExcelImportError:
    row: int
    column: str
    message: str

    def to_dict(self) -> dict:
        return {
            'row': self.row,
            'column': self.column,
            'message': self.message,
        }


def _generate_unique_nomenclature_code() -> str:
    code = f"nom_{uuid4().hex[:8]}"
    while NomenclatureItem.objects.filter(code=code).exists():
        code = f"nom_{uuid4().hex[:8]}"
    return code


def _parse_nomenclature_excel(file_obj) -> tuple[list[dict], list[ExcelImportError], dict]:
    """
    Parse and validate nomenclature import Excel.

    Returns:
        rows: list of parsed row dicts (including resolved FK ids/names when possible)
        errors: list of ExcelImportError across the whole file
        summary: counts
    """
    errors: list[ExcelImportError] = []
    parsed_rows: list[dict] = []
    seen_keys: set[tuple[str, str]] = set()  # (category_id, lower(name))

    wb = load_workbook(filename=file_obj, read_only=True, data_only=True)
    ws = wb.active

    total_data_rows = 0
    for excel_row_idx, row_cells in enumerate(ws.iter_rows(min_row=2, max_col=7, values_only=True), start=2):
        # Skip completely empty rows
        values = list(row_cells) if row_cells is not None else []
        if not values:
            continue

        # Normalize to length 7
        while len(values) < 7:
            values.append(None)

        if all((_to_str(v) == '') for v in values[:7]):
            continue

        total_data_rows += 1

        category_name = _to_str(values[0])
        name = _to_str(values[1])
        drawing_number = _to_str(values[2])
        unit = _to_str(values[3])
        description = _to_str(values[4])
        specifications = _to_str(values[5])
        nomenclature_type_name = _to_str(values[6])

        row_errors: list[ExcelImportError] = []

        category_obj = None
        if not category_name:
            row_errors.append(ExcelImportError(row=excel_row_idx, column='A', message='Не указан "Вид справочника" (колонка A).'))
        else:
            category_obj = CatalogCategory.objects.filter(is_active=True, name__iexact=category_name).first()
            if not category_obj:
                row_errors.append(
                    ExcelImportError(
                        row=excel_row_idx,
                        column='A',
                        message=f'Вид справочника "{category_name}" не найден в настройках справочников.'
                    )
                )

        if not name:
            row_errors.append(ExcelImportError(row=excel_row_idx, column='B', message='Не указано "Наименование" (колонка B).'))

        # Если в файле единица измерения не указана — подставим ниже:
        # 1) из типа номенклатуры (для закупаемых),
        # 2) иначе 'шт'.

        nomenclature_type_obj = None
        if category_obj and category_obj.is_purchased:
            active_types_qs = category_obj.nomenclature_types.filter(is_active=True)
            has_types = active_types_qs.exists()
            if nomenclature_type_name:
                nomenclature_type_obj = active_types_qs.filter(name__iexact=nomenclature_type_name).first()
                if not nomenclature_type_obj:
                    row_errors.append(
                        ExcelImportError(
                            row=excel_row_idx,
                            column='G',
                            message=f'Тип номенклатуры "{nomenclature_type_name}" не найден в виде справочника "{category_obj.name}".'
                        )
                    )
            else:
                if has_types:
                    row_errors.append(
                        ExcelImportError(
                            row=excel_row_idx,
                            column='G',
                            message=f'Для закупаемого вида "{category_obj.name}" нужно указать "Тип номенклатуры" (колонка G).'
                        )
                    )

        # Unit: if empty, try to take from nomenclature_type.default_unit, else 'шт'
        if not unit:
            if nomenclature_type_obj and getattr(nomenclature_type_obj, 'default_unit', None):
                unit = _to_str(nomenclature_type_obj.default_unit) or 'шт'
            else:
                unit = 'шт'

        # Manufactured: drawing number is optional (can be added later). For purchased, ignore drawing number.
        if category_obj and category_obj.is_purchased:
            drawing_number = ''

        # Duplicate checks
        if category_obj and name:
            key = (str(category_obj.id), name.lower())
            if key in seen_keys:
                row_errors.append(
                    ExcelImportError(
                        row=excel_row_idx,
                        column='B',
                        message='Дублирование в файле: такая позиция (вид справочника + наименование) уже встречалась.'
                    )
                )
            else:
                seen_keys.add(key)

            if NomenclatureItem.objects.filter(is_active=True, catalog_category=category_obj, name__iexact=name).exists():
                row_errors.append(
                    ExcelImportError(
                        row=excel_row_idx,
                        column='B',
                        message=f'Такая позиция уже существует в справочнике ("{category_obj.name}" / "{name}").'
                    )
                )

        errors.extend(row_errors)

        parsed_rows.append({
            'row': excel_row_idx,
            'catalog_category': str(category_obj.id) if category_obj else None,
            'catalog_category_name': category_obj.name if category_obj else category_name,
            'is_purchased': bool(category_obj.is_purchased) if category_obj else None,
            'name': name,
            'drawing_number': drawing_number,
            'unit': unit,
            'description': description,
            'specifications': specifications,
            'nomenclature_type': str(nomenclature_type_obj.id) if nomenclature_type_obj else None,
            'nomenclature_type_name': nomenclature_type_obj.name if nomenclature_type_obj else nomenclature_type_name,
            'can_import': len(row_errors) == 0,
            'row_errors': [e.to_dict() for e in row_errors],
        })

    summary = {
        'total_rows': total_data_rows,
        'parsed_rows': len(parsed_rows),
        'valid_rows': sum(1 for r in parsed_rows if r.get('can_import')),
        'error_rows': len({e.row for e in errors}),
        'errors_count': len(errors),
    }

    return parsed_rows, errors, summary


class NomenclatureFilterSet(django_filters.FilterSet):
    """Custom filterset for nomenclature items."""
    
    has_bom = django_filters.BooleanFilter(method='filter_has_bom')
    is_purchased = django_filters.BooleanFilter(field_name='catalog_category__is_purchased')
    unit = django_filters.CharFilter(field_name='unit', lookup_expr='iexact')
    primary_supplier = django_filters.UUIDFilter(method='filter_primary_supplier')
    nomenclature_type_isnull = django_filters.BooleanFilter(field_name='nomenclature_type', lookup_expr='isnull')
    
    class Meta:
        model = NomenclatureItem
        fields = [
            'catalog_category',
            'nomenclature_type',
            'nomenclature_type_isnull',
            'is_active',
            'has_bom',
            'is_purchased',
            'unit',
            'primary_supplier',
        ]
    
    def filter_has_bom(self, queryset, name, value):
        """Filter by whether item has a BOM structure.
        
        При фильтрации по BOM показываем только изготавливаемые изделия,
        т.к. закупаемые не имеют состава по определению.
        """
        # Фильтруем только изготавливаемые изделия
        queryset = queryset.filter(
            catalog_category__is_purchased=False
        )
        
        bom_exists = BOMStructure.objects.filter(
            root_item=OuterRef('pk'),
            is_active=True
        )
        if value:
            # has_bom=True: изготавливаемые с заполненным составом
            return queryset.annotate(
                has_bom_exists=Exists(bom_exists)
            ).filter(has_bom_exists=True)
        else:
            # has_bom=False: изготавливаемые без состава
            return queryset.annotate(
                has_bom_exists=Exists(bom_exists)
            ).filter(has_bom_exists=False)

    def filter_primary_supplier(self, queryset, name, value):
        """Filter by primary supplier (supplier id) for purchased items."""
        if not value:
            return queryset
        return queryset.filter(
            item_suppliers__is_primary=True,
            item_suppliers__supplier_id=value,
        ).distinct()


class CatalogCategoryViewSet(BaseModelViewSet):
    """
    ViewSet for catalog categories (nomenclature types).
    
    Endpoints:
    - GET /catalog-categories/ - list all categories
    - POST /catalog-categories/ - create category
    - GET /catalog-categories/{id}/ - get category details
    - PUT/PATCH /catalog-categories/{id}/ - update category
    - DELETE /catalog-categories/{id}/ - delete category
    - GET /catalog-categories/purchased/ - get purchased categories
    - GET /catalog-categories/manufactured/ - get manufactured categories
    """
    
    queryset = CatalogCategory.objects.all()
    
    serializer_classes = {
        'list': CatalogCategoryListSerializer,
        'retrieve': CatalogCategoryDetailSerializer,
        'default': CatalogCategoryDetailSerializer,
    }
    
    search_fields = ['name', 'code', 'description']
    filterset_fields = ['is_purchased', 'is_active']
    ordering_fields = ['sort_order', 'name']
    ordering = ['sort_order', 'name']
    
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    
    def get_serializer_class(self):
        return self.serializer_classes.get(
            self.action,
            self.serializer_classes['default']
        )
    
    @action(detail=False, methods=['get'])
    def purchased(self, request):
        """Get only purchased categories."""
        categories = self.get_queryset().filter(is_purchased=True, is_active=True)
        serializer = CatalogCategoryListSerializer(categories, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def manufactured(self, request):
        """Get only manufactured categories."""
        categories = self.get_queryset().filter(is_purchased=False, is_active=True)
        serializer = CatalogCategoryListSerializer(categories, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def allowed_children(self, request, pk=None):
        """Get allowed children for this category."""
        category = self.get_object()
        children = category.allowed_children.filter(is_active=True)
        serializer = CatalogCategoryListSerializer(children, many=True)
        return Response(serializer.data)
    
    def destroy(self, request, *args, **kwargs):
        """
        Удаление категории с каскадным удалением связанных данных.
        
        При удалении категории:
        1. Удаляются все типы номенклатуры (NomenclatureType)
        2. Удаляются все номенклатурные позиции (NomenclatureItem)
        """
        instance = self.get_object()
        
        # Получаем статистику для информирования
        types_count = instance.nomenclature_types.count()
        items_count = instance.nomenclature_items.count()
        
        # Удаляем номенклатурные позиции сначала (чтобы обойти PROTECT)
        instance.nomenclature_items.all().delete()
        
        # Удаляем типы номенклатуры (CASCADE сработает автоматически, но для ясности)
        instance.nomenclature_types.all().delete()
        
        # Удаляем саму категорию
        instance.delete()
        
        return Response(
            {
                'message': f'Категория удалена вместе с {types_count} типами и {items_count} номенклатурными позициями'
            },
            status=status.HTTP_200_OK
        )


class NomenclatureTypeViewSet(BaseModelViewSet):
    """
    ViewSet for nomenclature types.
    
    Endpoints:
    - GET /nomenclature-types/ - list all types
    - POST /nomenclature-types/ - create type
    - GET /nomenclature-types/{id}/ - get type details
    - PUT/PATCH /nomenclature-types/{id}/ - update type
    - DELETE /nomenclature-types/{id}/ - delete type
    """
    
    queryset = NomenclatureType.objects.select_related('catalog_category').all()
    serializer_class = NomenclatureTypeSerializer
    search_fields = ['name', 'description']
    filterset_fields = ['catalog_category', 'is_active']
    ordering_fields = ['catalog_category__sort_order', 'name']
    ordering = ['catalog_category__sort_order', 'name']
    
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]


class NomenclatureViewSet(BulkActionMixin, BaseModelViewSet):
    """
    ViewSet for nomenclature items.
    
    Endpoints:
    - GET /nomenclature/ - list all items
    - POST /nomenclature/ - create item
    - GET /nomenclature/{id}/ - get item details
    - PUT/PATCH /nomenclature/{id}/ - update item
    - DELETE /nomenclature/{id}/ - soft delete item
    - GET /nomenclature/tree/ - get hierarchical tree
    - GET /nomenclature/by-category/ - get items grouped by category
    - GET /nomenclature/categories/ - get category choices
    """
    
    queryset = NomenclatureItem.objects.select_related(
        'catalog_category', 'nomenclature_type'
    ).prefetch_related('item_suppliers__supplier').filter(is_active=True)
    
    serializer_classes = {
        'list': NomenclatureListSerializer,
        'retrieve': NomenclatureDetailSerializer,
        'default': NomenclatureDetailSerializer,
    }
    
    search_fields = ['name', 'drawing_number', 'description']
    filterset_class = NomenclatureFilterSet
    ordering_fields = [
        'name',
        'drawing_number',
        'unit',
        'created_at',
        'catalog_category__is_purchased',
        'catalog_category__sort_order',
        'catalog_category__name',
        'nomenclature_type__name',
    ]
    ordering = ['catalog_category__is_purchased', 'catalog_category__sort_order', 'name']
    
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    
    def get_serializer_class(self):
        return self.serializer_classes.get(
            self.action,
            self.serializer_classes['default']
        )
    
    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Get available nomenclature categories (catalog categories)."""
        categories = CatalogCategory.objects.filter(is_active=True).order_by('sort_order')
        serializer = CatalogCategoryListSerializer(categories, many=True)
        return Response(serializer.data)

    @action(
        detail=False,
        methods=['post'],
        url_path='import-excel/preview',
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel_preview(self, request):
        """Preview import of nomenclature items from Excel (.xlsx).

        Expects multipart/form-data with file field named `file`.
        Returns preview rows + all validation errors (with row/column).
        """
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': 'Не передан файл (поле "file").'}, status=status.HTTP_400_BAD_REQUEST)
        if not upload.name.lower().endswith('.xlsx'):
            return Response({'error': 'Поддерживается только формат .xlsx.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows, errors, summary = _parse_nomenclature_excel(upload)
        except Exception as exc:
            return Response(
                {'error': f'Не удалось прочитать Excel файл: {exc}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({
            'rows': rows,
            'errors': [e.to_dict() for e in errors],
            'summary': summary,
        })

    @action(
        detail=False,
        methods=['post'],
        url_path='import-excel/confirm',
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel_confirm(self, request):
        """Confirm import of nomenclature items from Excel (.xlsx).

        Re-reads and re-validates the Excel file. If any errors exist, nothing is created.
        """
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': 'Не передан файл (поле "file").'}, status=status.HTTP_400_BAD_REQUEST)
        if not upload.name.lower().endswith('.xlsx'):
            return Response({'error': 'Поддерживается только формат .xlsx.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows, errors, summary = _parse_nomenclature_excel(upload)
        except Exception as exc:
            return Response(
                {'error': f'Не удалось прочитать Excel файл: {exc}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if errors:
            return Response(
                {
                    'error': 'В файле есть ошибки. Исправьте их и повторите импорт.',
                    'rows': rows,
                    'errors': [e.to_dict() for e in errors],
                    'summary': summary,
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        created = 0
        created_ids: list[str] = []
        with transaction.atomic():
            for r in rows:
                if not r.get('can_import'):
                    continue
                category_id = r.get('catalog_category')
                if not category_id:
                    continue
                category_obj = CatalogCategory.objects.get(pk=category_id)
                nomenclature_type_obj = None
                if r.get('nomenclature_type'):
                    nomenclature_type_obj = NomenclatureType.objects.get(pk=r['nomenclature_type'])

                item = NomenclatureItem.objects.create(
                    code=_generate_unique_nomenclature_code(),
                    name=r.get('name') or '',
                    catalog_category=category_obj,
                    nomenclature_type=nomenclature_type_obj,
                    drawing_number=r.get('drawing_number') or '',
                    unit=r.get('unit') or 'шт',
                    description=r.get('description') or '',
                    specifications=r.get('specifications') or '',
                    created_by=request.user,
                    updated_by=request.user,
                )
                created += 1
                created_ids.append(str(item.id))

        return Response({
            'created': created,
            'created_ids': created_ids,
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=['get'])
    def legacy_categories(self, request):
        """Get legacy category choices (for backward compatibility)."""
        return Response(NomenclatureCategorySerializer.get_choices())
    
    @action(detail=False, methods=['get'])
    def tree(self, request):
        """Get nomenclature items as hierarchical tree by category."""
        catalog_category = request.query_params.get('catalog_category')
        
        queryset = self.get_queryset()
        if catalog_category:
            queryset = queryset.filter(catalog_category_id=catalog_category)
        
        # Group by catalog_category for tree structure
        categories = {}
        for item in queryset:
            cat_id = str(item.catalog_category_id) if item.catalog_category else 'uncategorized'
            cat_name = item.catalog_category.name if item.catalog_category else 'Без категории'
            if cat_id not in categories:
                categories[cat_id] = {
                    'id': cat_id,
                    'name': cat_name,
                    'items': []
                }
            categories[cat_id]['items'].append(
                NomenclatureListSerializer(item).data
            )
        
        return Response(list(categories.values()))
    
    @action(detail=False, methods=['get'])
    def by_category(self, request):
        """Get items grouped by catalog category with counts."""
        counts = self.get_queryset().values(
            'catalog_category', 'catalog_category__name'
        ).annotate(
            count=Count('id')
        ).order_by('catalog_category__sort_order')
        
        result = [
            {
                'catalog_category': item['catalog_category'],
                'name': item['catalog_category__name'] or 'Без категории',
                'count': item['count']
            }
            for item in counts
        ]
        
        return Response(result)
    
    @action(detail=False, methods=['get'])
    def search_advanced(self, request):
        """Advanced search with multiple filters."""
        queryset = self.get_queryset()
        
        # Apply filters
        catalog_category = request.query_params.get('catalog_category')
        supplier = request.query_params.get('supplier')
        nomenclature_type = request.query_params.get('type')
        
        if catalog_category:
            queryset = queryset.filter(catalog_category_id=catalog_category)
        if supplier:
            queryset = queryset.filter(item_suppliers__supplier_id=supplier)
        if nomenclature_type:
            queryset = queryset.filter(nomenclature_type_id=nomenclature_type)
        
        # Search
        search = request.query_params.get('q')
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(drawing_number__icontains=search)
            )
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = NomenclatureListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = NomenclatureListSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def usage(self, request, pk=None):
        """Get where this nomenclature is used (in BOMs, projects)."""
        item = self.get_object()
        
        # Find usage in BOMs
        bom_usage = item.bom_parents.select_related('bom').values(
            'bom__id', 'bom__name', 'quantity'
        )
        
        # Find usage in Projects
        project_usage = item.project_items.select_related('project').values(
            'project__id', 'project__name',
            'quantity', 'manufacturing_status'
        )
        
        return Response({
            'bom_usage': list(bom_usage),
            'project_usage': list(project_usage),
        })
    
    @action(detail=True, methods=['get', 'post', 'delete'])
    def suppliers(self, request, pk=None):
        """Manage suppliers for nomenclature item."""
        item = self.get_object()
        
        if request.method == 'GET':
            suppliers = item.item_suppliers.filter(is_active=True).select_related('supplier')
            serializer = NomenclatureSupplierSerializer(suppliers, many=True)
            return Response(serializer.data)
        
        elif request.method == 'POST':
            if not item.is_purchased:
                return Response(
                    {'error': 'Поставщиков можно указать только для закупаемых позиций'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            serializer = NomenclatureSupplierSerializer(data={
                **request.data,
                'nomenclature_item': item.id
            })
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.method == 'DELETE':
            supplier_id = request.data.get('supplier_id')
            if not supplier_id:
                return Response(
                    {'error': 'Необходимо указать supplier_id'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            deleted, _ = item.item_suppliers.filter(supplier_id=supplier_id).delete()
            if deleted:
                return Response(status=status.HTTP_204_NO_CONTENT)
            return Response(
                {'error': 'Связь не найдена'},
                status=status.HTTP_404_NOT_FOUND
            )


class SupplierViewSet(BulkActionMixin, BaseModelViewSet):
    """
    ViewSet for suppliers.
    
    Endpoints:
    - GET /suppliers/ - list all suppliers
    - POST /suppliers/ - create supplier
    - GET /suppliers/{id}/ - get supplier details
    - PUT/PATCH /suppliers/{id}/ - update supplier
    - DELETE /suppliers/{id}/ - delete supplier
    - GET /suppliers/top-rated/ - get top-rated suppliers
    - GET /suppliers/{id}/contacts/ - get contact persons
    """
    
    queryset = Supplier.objects.prefetch_related('contact_persons').filter(is_active=True)
    
    serializer_classes = {
        'list': SupplierListSerializer,
        'retrieve': SupplierDetailSerializer,
        'default': SupplierDetailSerializer,
    }
    
    search_fields = ['name', 'short_name', 'inn']
    filterset_fields = ['is_active', 'rating']
    ordering_fields = ['name', 'rating', 'created_at']
    ordering = ['name']
    
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    
    def get_serializer_class(self):
        return self.serializer_classes.get(
            self.action,
            self.serializer_classes['default']
        )
    
    @action(detail=False, methods=['get'])
    def top_rated(self, request):
        """Get top-rated suppliers."""
        limit = int(request.query_params.get('limit', 10))
        suppliers = self.get_queryset().filter(
            is_active=True,
            rating__isnull=False
        ).order_by('-rating')[:limit]
        
        serializer = SupplierListSerializer(suppliers, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def nomenclature(self, request, pk=None):
        """Get nomenclature items for this supplier."""
        supplier = self.get_object()
        items = NomenclatureItem.objects.filter(
            item_suppliers__supplier=supplier,
            is_active=True
        )
        serializer = NomenclatureListSerializer(items, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get', 'post', 'delete'])
    def contacts(self, request, pk=None):
        """Manage contact persons for supplier."""
        supplier = self.get_object()
        
        if request.method == 'GET':
            contacts = supplier.contact_persons.filter(is_active=True)
            serializer = ContactPersonSerializer(contacts, many=True)
            return Response(serializer.data)
        
        elif request.method == 'POST':
            serializer = ContactPersonCreateSerializer(data={
                **request.data,
                'supplier': supplier.id
            })
            if serializer.is_valid():
                contact = serializer.save()
                return Response(
                    ContactPersonSerializer(contact).data,
                    status=status.HTTP_201_CREATED
                )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.method == 'DELETE':
            contact_id = request.data.get('contact_id')
            if not contact_id:
                return Response(
                    {'error': 'Необходимо указать contact_id'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            deleted, _ = supplier.contact_persons.filter(id=contact_id).delete()
            if deleted:
                return Response(status=status.HTTP_204_NO_CONTENT)
            return Response(
                {'error': 'Контактное лицо не найдено'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['get', 'post', 'delete'])
    def bank_details(self, request, pk=None):
        """Manage bank details for supplier."""
        supplier = self.get_object()
        
        if request.method == 'GET':
            bank_details = supplier.bank_details.filter(is_active=True)
            serializer = BankDetailsInlineSerializer(bank_details, many=True)
            return Response(serializer.data)
        
        elif request.method == 'POST':
            serializer = BankDetailsSerializer(data={
                **request.data,
                'supplier': supplier.id
            })
            if serializer.is_valid():
                bank_detail = serializer.save()
                return Response(
                    BankDetailsSerializer(bank_detail).data,
                    status=status.HTTP_201_CREATED
                )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.method == 'DELETE':
            bank_detail_id = request.data.get('bank_detail_id')
            if not bank_detail_id:
                return Response(
                    {'error': 'Необходимо указать bank_detail_id'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            deleted, _ = supplier.bank_details.filter(id=bank_detail_id).delete()
            if deleted:
                return Response(status=status.HTTP_204_NO_CONTENT)
            return Response(
                {'error': 'Банковские реквизиты не найдены'},
                status=status.HTTP_404_NOT_FOUND
            )


class ContractorViewSet(BulkActionMixin, BaseModelViewSet):
    """
    ViewSet for contractors.
    
    Endpoints:
    - GET /contractors/ - list all contractors
    - POST /contractors/ - create contractor
    - GET /contractors/{id}/ - get contractor details
    - PUT/PATCH /contractors/{id}/ - update contractor
    - DELETE /contractors/{id}/ - delete contractor
    - GET /contractors/top-rated/ - get top-rated contractors
    - GET /contractors/{id}/contacts/ - get contact persons
    """
    
    queryset = Contractor.objects.prefetch_related('contact_persons').filter(is_active=True)
    
    serializer_classes = {
        'list': ContractorListSerializer,
        'retrieve': ContractorDetailSerializer,
        'default': ContractorDetailSerializer,
    }
    
    search_fields = ['name', 'short_name', 'inn', 'specialization']
    filterset_fields = ['is_active', 'rating']
    ordering_fields = ['name', 'rating', 'created_at']
    ordering = ['name']
    
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    
    def get_serializer_class(self):
        return self.serializer_classes.get(
            self.action,
            self.serializer_classes['default']
        )
    
    @action(detail=False, methods=['get'])
    def top_rated(self, request):
        """Get top-rated contractors."""
        limit = int(request.query_params.get('limit', 10))
        contractors = self.get_queryset().filter(
            is_active=True,
            rating__isnull=False
        ).order_by('-rating')[:limit]
        
        serializer = ContractorListSerializer(contractors, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def by_specialization(self, request):
        """Filter contractors by specialization."""
        specialization = request.query_params.get('q')
        if not specialization:
            return Response(
                {'error': 'Необходимо указать параметр q'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        contractors = self.get_queryset().filter(
            specialization__icontains=specialization
        )
        
        serializer = ContractorListSerializer(contractors, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get', 'post', 'delete'])
    def contacts(self, request, pk=None):
        """Manage contact persons for contractor."""
        contractor = self.get_object()
        
        if request.method == 'GET':
            contacts = contractor.contact_persons.filter(is_active=True)
            serializer = ContactPersonSerializer(contacts, many=True)
            return Response(serializer.data)
        
        elif request.method == 'POST':
            serializer = ContactPersonCreateSerializer(data={
                **request.data,
                'contractor': contractor.id
            })
            if serializer.is_valid():
                contact = serializer.save()
                return Response(
                    ContactPersonSerializer(contact).data,
                    status=status.HTTP_201_CREATED
                )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.method == 'DELETE':
            contact_id = request.data.get('contact_id')
            if not contact_id:
                return Response(
                    {'error': 'Необходимо указать contact_id'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            deleted, _ = contractor.contact_persons.filter(id=contact_id).delete()
            if deleted:
                return Response(status=status.HTTP_204_NO_CONTENT)
            return Response(
                {'error': 'Контактное лицо не найдено'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['get', 'post', 'delete'])
    def bank_details(self, request, pk=None):
        """Manage bank details for contractor."""
        contractor = self.get_object()
        
        if request.method == 'GET':
            bank_details = contractor.bank_details.filter(is_active=True)
            serializer = BankDetailsInlineSerializer(bank_details, many=True)
            return Response(serializer.data)
        
        elif request.method == 'POST':
            serializer = BankDetailsSerializer(data={
                **request.data,
                'contractor': contractor.id
            })
            if serializer.is_valid():
                bank_detail = serializer.save()
                return Response(
                    BankDetailsSerializer(bank_detail).data,
                    status=status.HTTP_201_CREATED
                )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        elif request.method == 'DELETE':
            bank_detail_id = request.data.get('bank_detail_id')
            if not bank_detail_id:
                return Response(
                    {'error': 'Необходимо указать bank_detail_id'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            deleted, _ = contractor.bank_details.filter(id=bank_detail_id).delete()
            if deleted:
                return Response(status=status.HTTP_204_NO_CONTENT)
            return Response(
                {'error': 'Банковские реквизиты не найдены'},
                status=status.HTTP_404_NOT_FOUND
            )


class ContactPersonViewSet(BaseModelViewSet):
    """
    ViewSet for contact persons.
    
    Endpoints:
    - GET /contact-persons/ - list all contact persons
    - POST /contact-persons/ - create contact person
    - GET /contact-persons/{id}/ - get contact person details
    - PUT/PATCH /contact-persons/{id}/ - update contact person
    - DELETE /contact-persons/{id}/ - delete contact person
    """
    
    queryset = ContactPerson.objects.select_related('supplier', 'contractor').filter(is_active=True)
    serializer_class = ContactPersonSerializer
    
    search_fields = ['last_name', 'first_name', 'phone', 'email']
    filterset_fields = ['supplier', 'contractor', 'is_primary', 'is_active']
    ordering_fields = ['last_name', 'is_primary']
    ordering = ['-is_primary', 'last_name']
    
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]


class DelayReasonViewSet(BaseModelViewSet):
    """
    ViewSet for delay reasons.
    
    Endpoints:
    - GET /delay-reasons/ - list all delay reasons
    - POST /delay-reasons/ - create delay reason
    - GET /delay-reasons/{id}/ - get delay reason details
    - PUT/PATCH /delay-reasons/{id}/ - update delay reason
    - DELETE /delay-reasons/{id}/ - delete delay reason
    """
    
    queryset = DelayReason.objects.filter(is_active=True)
    serializer_class = DelayReasonSerializer
    
    search_fields = ['name', 'description']
    filterset_fields = ['applies_to_procurement', 'applies_to_production', 'production_config', 'is_active']
    ordering_fields = ['name']
    ordering = ['name']
    
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    
    def get_queryset(self):
        """Filter delay reasons by production configuration."""
        queryset = super().get_queryset()
        
        # Get production configuration filter
        production_config = self.request.query_params.get('production_config')
        
        if production_config and production_config != 'all':
            # Return reasons that match this config or 'all'
            queryset = queryset.filter(
                models.Q(production_config=production_config) | 
                models.Q(production_config='all')
            )
        
        return queryset


class NomenclatureSupplierViewSet(BaseModelViewSet):
    """
    ViewSet for nomenclature-supplier links.
    
    Endpoints:
    - GET /nomenclature-suppliers/ - list all links
    - POST /nomenclature-suppliers/ - create link
    - GET /nomenclature-suppliers/{id}/ - get link details
    - PUT/PATCH /nomenclature-suppliers/{id}/ - update link
    - DELETE /nomenclature-suppliers/{id}/ - delete link
    """
    
    queryset = NomenclatureSupplier.objects.select_related(
        'nomenclature_item', 'supplier'
    ).filter(is_active=True)
    serializer_class = NomenclatureSupplierSerializer
    
    filterset_fields = ['nomenclature_item', 'supplier', 'is_primary', 'is_active']
    ordering_fields = ['is_primary', 'delivery_days']
    ordering = ['-is_primary', 'delivery_days']
    
    filter_backends = [
        DjangoFilterBackend,
        filters.OrderingFilter,
    ]


class BankDetailsViewSet(BaseModelViewSet):
    """
    ViewSet for bank details.
    
    Endpoints:
    - GET /bank-details/ - list all bank details
    - POST /bank-details/ - create bank detail
    - GET /bank-details/{id}/ - get bank detail details
    - PUT/PATCH /bank-details/{id}/ - update bank detail
    - DELETE /bank-details/{id}/ - delete bank detail
    - GET /bank-details/by-supplier/{supplier_id}/ - get bank details for supplier
    - GET /bank-details/by-contractor/{contractor_id}/ - get bank details for contractor
    """
    
    queryset = BankDetails.objects.select_related('supplier', 'contractor').filter(is_active=True)
    serializer_class = BankDetailsSerializer
    
    filterset_fields = ['supplier', 'contractor', 'is_primary', 'is_active', 'currency']
    search_fields = ['bank_name', 'bik', 'settlement_account']
    ordering_fields = ['bank_name', 'is_primary']
    ordering = ['-is_primary', 'bank_name']
    
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    
    @action(detail=False, methods=['get'], url_path='by-supplier/(?P<supplier_id>[^/.]+)')
    def by_supplier(self, request, supplier_id=None):
        """Get all bank details for a specific supplier."""
        bank_details = self.get_queryset().filter(supplier_id=supplier_id)
        serializer = self.get_serializer(bank_details, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='by-contractor/(?P<contractor_id>[^/.]+)')
    def by_contractor(self, request, contractor_id=None):
        """Get all bank details for a specific contractor."""
        bank_details = self.get_queryset().filter(contractor_id=contractor_id)
        serializer = self.get_serializer(bank_details, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def set_primary(self, request, pk=None):
        """Set this bank detail as primary."""
        bank_detail = self.get_object()
        
        # Reset other primary bank details for the same entity
        if bank_detail.supplier:
            BankDetails.objects.filter(supplier=bank_detail.supplier, is_primary=True).update(is_primary=False)
        elif bank_detail.contractor:
            BankDetails.objects.filter(contractor=bank_detail.contractor, is_primary=True).update(is_primary=False)
        
        bank_detail.is_primary = True
        bank_detail.save()
        
        serializer = self.get_serializer(bank_detail)
        return Response(serializer.data)
